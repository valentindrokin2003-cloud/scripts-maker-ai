import openpyxl
import pytest

from src.brief_extractor import _extract_product_words_from_excel_text
from src.excel_parser import parse_excel_to_text
from src.words_matcher import match_words


REAL_DICT_PATH = "data/words_ok_groups_v2.xlsx"


@pytest.fixture
def dict_xlsx(tmp_path):
    path = tmp_path / "words.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Группа1"
    ws["A1"] = "Панель"
    ws["A2"] = "Фасад"
    ws["A3"] = "Кассета"
    ws["A4"] = "Труба"
    wb.save(path)
    return str(path)


def test_exact_match(dict_xlsx):
    result = match_words(["Панель", "Труба"], dict_xlsx)
    assert "Панель" in result
    assert "Труба" in result


def test_fuzzy_match(dict_xlsx):
    result = match_words(["Панели"], dict_xlsx)
    assert "Панель" in result


def test_no_match_returns_empty(dict_xlsx):
    result = match_words(["Экскаватор"], dict_xlsx)
    assert result == []


def test_returns_dict_words_not_input_words(dict_xlsx):
    result = match_words(["Фасадные"], dict_xlsx)
    assert "Фасад" in result
    assert "Фасадные" not in result


def test_deduplication(dict_xlsx):
    result = match_words(["Панель", "Панели"], dict_xlsx)
    assert result.count("Панель") == 1


def test_uses_only_word_column_not_group_name(tmp_path):
    path = tmp_path / "words.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("word", "group_name", "group_id", "word_id"))
    ws.append(("Коннектор", "Детали и материалы для промышленности и техники", 1, 1))
    wb.save(path)

    result = match_words(["детали для промышленности"], str(path))

    assert result == []


def test_bzs_product_phrases_keep_only_target_product_words(tmp_path):
    path = tmp_path / "words.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("word", "group_name", "group_id", "word_id"))
    ws.append(("Разъем", "Электриеские и измерительные приборы", 24, 582))
    ws.append(("Уголок", "Металлопрокатная продукция", 61, 676))
    ws.append(("Шина медная", "Электротовары", 40, 897))
    ws.append(("Нить", "Мешки, скотч, товары для уборки", 68, 987))
    ws.append(("Стеклопакет", "Окные и дверные конструкции", 45, 1227))
    ws.append(("Провод", "Электротовары", 40, 1412))
    ws.append(("Коннектор", "Электротовары", 40, 1980))
    ws.append(("Светильник", "Электротовары", 40, 2834))
    wb.save(path)

    product_words = [
        "коннекторы для обогрева",
        "гермовводы",
        "герметичные соединения проводов",
        "коннекторы светильников",
        "уголок для сборки стеклопакетов с проводным соединением",
        "вольфрамовая нить",
        "медная луженая шинка (шина)",
        "промышленные разъемы",
    ]

    result = match_words(product_words, str(path))

    assert result == ["Коннектор", "Уголок", "Разъем"]


def test_wolfram_thread_does_not_match_cleaning_thread(tmp_path):
    path = tmp_path / "words.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("word", "group_name", "group_id", "word_id"))
    ws.append(("Нить", "Мешки, скотч, товары для уборки", 68, 987))
    wb.save(path)

    result = match_words(["вольфрамовая нить"], str(path))

    assert result == []


def _match_real_brief(path):
    excel_text = parse_excel_to_text(path)
    product_words = _extract_product_words_from_excel_text(excel_text)
    return match_words(product_words, REAL_DICT_PATH)


def test_bzs_real_brief_regression():
    assert _match_real_brief("data/ТЗ_BZS.xlsx") == [
        "Коннектор",
        "Уголок",
        "Разъем",
    ]


def test_oxrana_real_brief_does_not_match_poster():
    result = _match_real_brief("data/ТЗ_Oxrana.xlsx")

    assert result == ["Охрана помещения"]
    assert "Постер" not in result


def test_stomprom_real_brief_does_not_match_photo_wallpaper():
    result = _match_real_brief("data/ТЗ_Stomprom.xlsx")

    assert result == ["Зеркало", "Очки защитные", "Стоматологический материал"]
    assert "Фотообои" not in result


def test_timtrans_real_brief_does_not_match_connection_services():
    result = _match_real_brief("data/ТЗ_TimTrans.xlsx")

    assert result == ["Услуга доставки", "Перевозка груза"]
    assert "Услуга подключения" not in result


def test_bam_real_brief_matches_only_dictionary_food_item():
    assert _match_real_brief("data/ТЗ_BAM.xlsx") == ["Паста"]


def test_rufas_real_brief_regression():
    assert _match_real_brief("data/ТЗ_Rufas.xlsx") == [
        "Фасад",
        "Алюминий",
        "Панель",
    ]


def test_leodata_real_brief_regression():
    assert _match_real_brief("data/ТЗ_Leodata.xlsx") == ["Сервер"]


def test_transservis_real_brief_has_no_dictionary_false_positive():
    assert _match_real_brief("data/ТЗ_Transservis.xlsx") == []
