import nbformat
import openpyxl
import pytest

from src.brief_extractor import BriefExtractionError
from src.models import BriefData, BriefIssue, BriefReview
from src.brief_review import review_brief
from src.notebook_replacements import MARKER_TEMPLATES
from src.pipeline import run_pipeline
from src.settings import AgentSettings


def _write_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Бриф"
    ws.append(("Название компании заказчика", "ООО Фреш"))
    ws.append(("ИНН заказчика", "5009138436"))
    wb.save(path)


def _write_dict(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Слова"
    ws.append(("Панель",))
    ws.append(("Фасад",))
    wb.save(path)


def _write_template(path):
    nb = nbformat.v4.new_notebook()
    nb.cells = [
        nbformat.v4.new_code_cell(f"# {marker}\npass")
        for marker in MARKER_TEMPLATES
    ]
    nbformat.write(nb, path)


def _copy_block_templates(target_dir):
    source_dir = AgentSettings(api_key=None).notebook_blocks_dir
    target_dir.mkdir()
    for filename in MARKER_TEMPLATES.values():
        source = f"{source_dir}/{filename}"
        target = target_dir / filename
        with open(source, encoding="utf-8") as src:
            target.write_text(src.read(), encoding="utf-8")


def _settings(tmp_path):
    brief_path = tmp_path / "brief.xlsx"
    dict_path = tmp_path / "dict.xlsx"
    template_path = tmp_path / "template.ipynb"
    blocks_dir = tmp_path / "blocks"

    _write_xlsx(brief_path)
    _write_dict(dict_path)
    _write_template(template_path)
    _copy_block_templates(blocks_dir)

    return brief_path, AgentSettings(
        api_key=None,
        template_path=str(template_path),
        notebook_blocks_dir=str(blocks_dir),
        dict_path=str(dict_path),
        output_dir=str(tmp_path / "output"),
    )


def _extract_brief(excel_text, client, model):
    assert "ООО Фреш" in excel_text
    return BriefData(
        name="ООО Фреш",
        inn_client=["5009138436"],
        analysis_period="range:2025-01-01:2025-01-31",
        product_words=["Панель"],
    )


def _generate_regex(product_words, client, model):
    assert product_words == ["Панель"]
    return [r"\bпанел\w{0,3}\b"]


def _review_ok(excel_text, brief):
    return BriefReview(
        status="ok",
        issues=[],
        extracted_fields={"client_name": brief.name},
    )


def test_run_pipeline_generates_notebook(tmp_path):
    brief_path, settings = _settings(tmp_path)

    result = run_pipeline(
        str(brief_path),
        settings.output_dir,
        settings,
        client=None,
        brief_extractor=_extract_brief,
        brief_reviewer=_review_ok,
        regex_generator=_generate_regex,
    )

    assert result.status == "completed"
    assert result.output_path.endswith("ООО_Фреш_script.ipynb")
    assert result.start_date == "2025-01-01"
    assert result.end_date == "2025-01-31"
    assert result.lst_sbersov == ["Панель"]
    assert result.list_words == [r"\bпанел\w{0,3}\b"]

    nb = nbformat.read(result.output_path, as_version=4)
    notebook_text = "\n".join(cell.source for cell in nb.cells)
    assert "name = 'ООО Фреш'" in notebook_text
    assert "list_words = [" in notebook_text


def test_run_pipeline_propagates_extraction_error(tmp_path):
    brief_path, settings = _settings(tmp_path)

    def fail_extract(excel_text, client, model):
        raise BriefExtractionError("bad brief")

    with pytest.raises(BriefExtractionError, match="bad brief"):
        run_pipeline(
            str(brief_path),
            settings.output_dir,
            settings,
            client=None,
            brief_extractor=fail_extract,
            brief_reviewer=_review_ok,
            regex_generator=_generate_regex,
        )


def test_run_pipeline_stops_when_brief_needs_revision(tmp_path):
    brief_path, settings = _settings(tmp_path)

    def review_needs_revision(excel_text, brief):
        return BriefReview(
            status="needs_revision",
            issues=[
                BriefIssue(
                    severity="warning",
                    title="Период неясен",
                    detail="Нужно уточнить период анализа.",
                    field_name="analysis_period",
                )
            ],
            extracted_fields={"client_name": brief.name, "analysis_period": brief.analysis_period},
        )

    result = run_pipeline(
        str(brief_path),
        settings.output_dir,
        settings,
        client=None,
        brief_extractor=_extract_brief,
        brief_reviewer=review_needs_revision,
        regex_generator=_generate_regex,
    )

    assert result.status == "needs_revision"
    assert result.output_path is None
    assert result.review is not None
    assert result.review.issues[0].title == "Период неясен"


def test_run_pipeline_continues_when_okved_conflicts_with_non_industry_brief(tmp_path):
    brief_path = tmp_path / "brief.xlsx"
    dict_path = tmp_path / "dict.xlsx"
    template_path = tmp_path / "template.ipynb"
    blocks_dir = tmp_path / "blocks"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Бриф"
    ws.append(("Название компании заказчика", "ООО Фреш"))
    ws.append(("ИНН заказчика", "5009138436"))
    ws.append(("Целевая аудитория", "любые потребители указанной продукции"))
    wb.save(brief_path)

    _write_dict(dict_path)
    _write_template(template_path)
    _copy_block_templates(blocks_dir)

    settings = AgentSettings(
        api_key=None,
        template_path=str(template_path),
        notebook_blocks_dir=str(blocks_dir),
        dict_path=str(dict_path),
        output_dir=str(tmp_path / "output"),
    )

    def extract_with_okved(excel_text, client, model):
        return BriefData(
            name="ООО Фреш",
            inn_client=["5009138436"],
            analysis_period="range:2025-01-01:2025-01-31",
            product_words=["Панель"],
            okved_list=["43.21"],
        )

    result = run_pipeline(
        str(brief_path),
        settings.output_dir,
        settings,
        client=None,
        brief_extractor=extract_with_okved,
        brief_reviewer=review_brief,
        regex_generator=_generate_regex,
    )

    assert result.status == "completed"
    assert result.output_path is not None
    assert result.review is not None
    assert result.review.status == "ok"
    okved_issue = next(
        issue for issue in result.review.issues
        if issue.title == "ОКВЭД-фильтр противоречит формулировке брифа"
    )
    assert okved_issue.severity == "recommendation"
