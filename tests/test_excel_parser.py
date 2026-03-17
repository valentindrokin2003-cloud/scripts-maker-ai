import openpyxl
import pytest
import os
from src.excel_parser import parse_excel_to_text


@pytest.fixture
def simple_xlsx(tmp_path):
    path = tmp_path / "brief.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws["A1"] = "Название компании заказчика"
    ws["B1"] = "ООО Фреш"
    ws["A2"] = "ИНН заказчика"
    ws["B2"] = "5009138436"
    wb.save(path)
    return str(path)


def test_parse_excel_returns_string(simple_xlsx):
    result = parse_excel_to_text(simple_xlsx)
    assert isinstance(result, str)


def test_parse_excel_contains_values(simple_xlsx):
    result = parse_excel_to_text(simple_xlsx)
    assert "ООО Фреш" in result
    assert "5009138436" in result
    assert "Название компании заказчика" in result


def test_parse_excel_includes_sheet_name(simple_xlsx):
    result = parse_excel_to_text(simple_xlsx)
    assert "Лист1" in result
