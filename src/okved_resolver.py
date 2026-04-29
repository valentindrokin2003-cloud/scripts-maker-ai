import logging
import os
import re
from dataclasses import dataclass
from functools import lru_cache
from typing import Any, Iterable

import openpyxl

from src.llm_json import LLMJsonError, parse_llm_json

logger = logging.getLogger(__name__)

DEFAULT_OKVED_PATH = "data/Kod_Okved-2.xlsx"

STOP_OKVED_PATTERNS = (
    re.compile(r"\b(?:люб\w+|все)\s+(?:потребител\w+|клиент\w+|компан\w+|организац\w+)\b", re.IGNORECASE),
    re.compile(r"\bбез\s+ограничени\w+\s+по\s+(?:оквэд|отрасл\w+)\b", re.IGNORECASE),
    re.compile(r"\bвсе\s+оквэд\b", re.IGNORECASE),
    re.compile(r"\bвсе\s+отрасл\w+\b", re.IGNORECASE),
)

EXPLICIT_OKVED_RE = re.compile(r"(?<!\d)(\d{2}(?:\.\d{1,2}){0,2})(?!\d)")

TARGET_LABEL_KEYWORDS = (
    "оквэд",
    "отрасл",
    "сфера",
    "сегмент",
    "потребител",
    "покупател",
    "клиент",
    "целевая аудитория",
    "целевая группа",
    "тип клиент",
    "вид деятельност",
)

NEGATIVE_CONTEXT_TERMS = {
    "заводы",
    "завод",
    "компании",
    "компания",
    "организации",
    "организация",
    "бизнес",
    "госкомпании",
    "государственные компании",
}

STOPWORDS = {
    "и",
    "или",
    "для",
    "по",
    "на",
    "с",
    "из",
    "под",
    "над",
    "без",
    "не",
    "все",
    "вся",
    "всех",
    "любые",
    "любой",
    "любая",
    "любых",
    "указанной",
    "указанные",
    "указанный",
    "продукции",
    "товаров",
    "товары",
    "покупатели",
    "потребители",
    "клиенты",
    "бизнес",
    "типа",
    "крупные",
    "государственные",
    "компании",
    "организации",
}

TOKEN_ENDINGS = (
    "ению",
    "анию",
    "иями",
    "ями",
    "ами",
    "ого",
    "его",
    "ому",
    "ему",
    "ыми",
    "ими",
    "ыми",
    "ыми",
    "ией",
    "ией",
    "ией",
    "ов",
    "ев",
    "ей",
    "ам",
    "ям",
    "ах",
    "ях",
    "ом",
    "ем",
    "ой",
    "ый",
    "ий",
    "ое",
    "ее",
    "ая",
    "яя",
    "ые",
    "ие",
    "ым",
    "им",
    "ую",
    "юю",
    "ию",
    "ью",
    "а",
    "я",
    "ы",
    "и",
    "е",
    "у",
    "ю",
)

MANUAL_CODE_HINTS = {
    "ук": ["68.32.1", "68.32.2"],
    "управляющая компания": ["68.32.1", "68.32.2"],
    "управляющие компании": ["68.32.1", "68.32.2"],
    "жкх": ["68.32.1", "68.32.2"],
    "застройщик": ["41.20", "71.12.2"],
    "застройщики": ["41.20", "71.12.2"],
    "девелопер": ["41.20", "71.12.2"],
    "девелоперы": ["41.20", "71.12.2"],
    "монтажная организация": ["43.21", "43.29"],
    "монтажные организации": ["43.21", "43.29"],
    "строительно монтажные организации": ["43.21", "43.29"],
    "санаторий": ["86.90.4"],
    "санатории": ["86.90.4"],
    "санаторно курортные организации": ["86.90.4"],
    "гостиницы": ["55.10"],
    "гостиница": ["55.10"],
}

RERANK_SYSTEM_PROMPT = """Ты выбираешь коды ОКВЭД для фильтрации B2B-выборки.
Тебе дают текст брифа и список кандидатов, уже поднятых из справочника.

Правила:
1. Выбирай ТОЛЬКО из списка кандидатов.
2. Если отраслевой фильтр не нужен, верни [].
3. Если в тексте сказано про любых / всех потребителей без отраслевого ограничения, верни [].
4. Не добавляй коды, которых нет в списке кандидатов.
5. Верни ТОЛЬКО JSON-массив кодов, например ["43.21", "71.12.2"].
"""


@dataclass(frozen=True)
class OkvedEntry:
    code: str
    title: str
    normalized_title: str
    tokens: frozenset[str]


@dataclass(frozen=True)
class OkvedResolution:
    codes: list[str]
    explanations: dict[str, list[str]]
    decision_reason: str


def _normalize_text(value: str) -> str:
    normalized = value.casefold().replace("ё", "е")
    normalized = normalized.replace("«", " ").replace("»", " ")
    normalized = re.sub(r"[^a-zа-я0-9.]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def _tokenize(value: str) -> list[str]:
    return re.findall(r"[a-zа-я0-9]+", _normalize_text(value))


def _canonical_token(token: str) -> str:
    for ending in TOKEN_ENDINGS:
        if token.endswith(ending) and len(token) - len(ending) >= 3:
            return token[: -len(ending)]
    return token


def _meaningful_tokens(value: str) -> set[str]:
    return {
        _canonical_token(token)
        for token in _tokenize(value)
        if token not in STOPWORDS and token not in NEGATIVE_CONTEXT_TERMS and len(token) >= 4
    }


def _is_target_label(label: str) -> bool:
    normalized = _normalize_text(label)
    if "продук" in normalized or "товар" in normalized:
        return False
    return any(keyword in normalized for keyword in TARGET_LABEL_KEYWORDS)


def _iter_labeled_rows(excel_text: str) -> Iterable[tuple[str, str]]:
    for line in excel_text.splitlines():
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < 2:
            continue
        label = _normalize_text(parts[0])
        value = " | ".join(part.strip() for part in parts[1:] if part.strip())
        if value:
            yield label, value


def _load_reference_path() -> str:
    return os.getenv("OKVED_DICT_PATH", DEFAULT_OKVED_PATH)


@lru_cache(maxsize=4)
def load_okved_reference(path: str) -> dict[str, OkvedEntry]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    code_index = None
    title_index = None
    for row in worksheet.iter_rows(values_only=True):
        normalized = [
            _normalize_text(str(cell)) if cell is not None else ""
            for cell in row
        ]
        if "код" in normalized and "название" in normalized:
            code_index = normalized.index("код")
            title_index = normalized.index("название")
            break

    if code_index is None or title_index is None:
        raise ValueError(f"Could not find OKVED header row in {path}")

    entries: dict[str, OkvedEntry] = {}
    for row in worksheet.iter_rows(values_only=True):
        if code_index >= len(row) or title_index >= len(row):
            continue
        code = row[code_index]
        title = row[title_index]
        if code is None or title is None:
            continue
        code_text = str(code).strip()
        title_text = str(title).strip()
        if not code_text or not title_text:
            continue
        if not re.fullmatch(r"[A-Z]|\d{2}(?:\.\d{1,2}){0,2}", code_text):
            continue
        entries[code_text] = OkvedEntry(
            code=code_text,
            title=title_text,
            normalized_title=_normalize_text(title_text),
            tokens=frozenset(_meaningful_tokens(title_text)),
        )

    return entries


def _extract_explicit_codes(excel_text: str, llm_codes: list[str]) -> list[str]:
    found = list(EXPLICIT_OKVED_RE.findall(excel_text))
    for value in llm_codes:
        if not isinstance(value, str):
            continue
        found.extend(EXPLICIT_OKVED_RE.findall(value))
    return list(dict.fromkeys(found))


def _has_stop_signal(excel_text: str) -> bool:
    return any(pattern.search(excel_text) for pattern in STOP_OKVED_PATTERNS)


def _collect_target_context(excel_text: str) -> str:
    contexts = [
        value
        for label, value in _iter_labeled_rows(excel_text)
        if _is_target_label(label)
    ]
    if contexts:
        return "\n".join(contexts)
    return excel_text


def has_okved_stop_signal(excel_text: str) -> bool:
    return _has_stop_signal(excel_text)


def has_okved_target_context(excel_text: str) -> bool:
    for label, value in _iter_labeled_rows(excel_text):
        if _is_target_label(label):
            return True
        if _has_manual_hint(value):
            return True

    normalized_text = _normalize_text(excel_text)
    if "оквэд" in normalized_text and _extract_explicit_codes(excel_text, []):
        return True
    return _has_manual_hint(excel_text)


def is_okved_code_too_broad(code: str) -> bool:
    normalized = str(code).strip()
    return bool(re.fullmatch(r"[A-Z]|\d{2}", normalized))


def _codes_from_manual_hints(context: str, valid_codes: set[str]) -> list[str]:
    normalized = _normalize_text(context)
    found: list[str] = []
    for phrase, codes in MANUAL_CODE_HINTS.items():
        pattern = rf"(?<![a-zа-я0-9]){re.escape(phrase)}(?![a-zа-я0-9])"
        if not re.search(pattern, normalized):
            continue
        for code in codes:
            if code in valid_codes and code not in found:
                found.append(code)
    return found


def _has_manual_hint(value: str) -> bool:
    normalized = _normalize_text(value)
    return any(
        re.search(rf"(?<![a-zа-я0-9]){re.escape(phrase)}(?![a-zа-я0-9])", normalized)
        for phrase in MANUAL_CODE_HINTS
    )


def _codes_from_semantic_search(context: str, reference: dict[str, OkvedEntry]) -> list[str]:
    normalized = _normalize_text(context)
    context_tokens = _meaningful_tokens(context)
    if not context_tokens:
        return []

    scored: list[tuple[int, str]] = []
    for entry in reference.values():
        if len(entry.code) < 4:
            continue
        score = 0
        if entry.normalized_title in normalized:
            score += 10
        overlap = context_tokens & entry.tokens
        score += len(overlap) * 2
        if score >= 4:
            scored.append((score, entry.code))

    scored.sort(key=lambda item: (-item[0], item[1]))
    result: list[str] = []
    for _, code in scored:
        if code not in result:
            result.append(code)
        if len(result) >= 6:
            break
    return result


def _filter_valid_codes(codes: list[str], reference: dict[str, OkvedEntry]) -> list[str]:
    return [code for code in codes if code in reference]


def _dedupe_codes(codes: list[str]) -> list[str]:
    return list(dict.fromkeys(codes))


def _append_explanation(
    explanations: dict[str, list[str]],
    code: str,
    reason: str,
) -> None:
    bucket = explanations.setdefault(code, [])
    if reason not in bucket:
        bucket.append(reason)


def _build_candidate_resolution(
    context: str,
    reference: dict[str, OkvedEntry],
) -> tuple[list[str], dict[str, list[str]]]:
    explanations: dict[str, list[str]] = {}

    normalized_context = _normalize_text(context)
    manual_codes: list[str] = []
    for phrase, codes in MANUAL_CODE_HINTS.items():
        pattern = rf"(?<![a-zа-я0-9]){re.escape(phrase)}(?![a-zа-я0-9])"
        if not re.search(pattern, normalized_context):
            continue
        for code in codes:
            if code in reference:
                if code not in manual_codes:
                    manual_codes.append(code)
                _append_explanation(
                    explanations,
                    code,
                    f"Семантический сигнал из брифа: '{phrase}'",
                )

    context_tokens = _meaningful_tokens(context)
    for entry in reference.values():
        if len(entry.code) < 4:
            continue
        overlap = sorted(context_tokens & entry.tokens)
        score = 0
        if entry.normalized_title in normalized_context:
            score += 10
            _append_explanation(
                explanations,
                entry.code,
                f"Точное совпадение с формулировкой ОКВЭД: {entry.title}",
            )
        if overlap:
            score += len(overlap) * 2
            _append_explanation(
                explanations,
                entry.code,
                f"Совпадение по смысловым токенам: {', '.join(overlap)}",
            )
        if score < 4 and entry.code not in manual_codes:
            explanations.pop(entry.code, None)

    candidate_codes = _dedupe_codes(manual_codes + list(explanations))
    limited_codes = candidate_codes[:12]
    return limited_codes, {code: explanations[code] for code in limited_codes}


def _llm_rerank_candidates(
    context: str,
    candidate_codes: list[str],
    reference: dict[str, OkvedEntry],
    client: Any,
    model: str,
) -> list[str] | None:
    candidate_lines = [
        f"- {code}: {reference[code].title}"
        for code in candidate_codes
        if code in reference
    ]
    if not candidate_lines:
        return []

    user_prompt = (
        "Текст брифа:\n"
        f"{context}\n\n"
        "Кандидаты ОКВЭД:\n"
        + "\n".join(candidate_lines)
    )

    try:
        message = client.chat.completions.create(
            model=model,
            max_tokens=512,
            timeout=45,
            messages=[
                {"role": "system", "content": RERANK_SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
        )
        raw = message.choices[0].message.content.strip()
        parsed = parse_llm_json(raw)
    except (LLMJsonError, AttributeError, IndexError, KeyError, TypeError, ValueError) as exc:
        logger.warning("[resolve_okved_list] OKVED rerank parse failed: %s", exc)
        return None
    except Exception as exc:  # noqa: BLE001 - selection fallback is intentional
        logger.warning("[resolve_okved_list] OKVED rerank request failed: %s: %s", type(exc).__name__, exc)
        return None

    if not isinstance(parsed, list):
        logger.warning("[resolve_okved_list] OKVED rerank returned %s instead of list", type(parsed).__name__)
        return None

    valid = [
        code
        for code in parsed
        if isinstance(code, str) and code in candidate_codes
    ]
    return _dedupe_codes(valid)


def resolve_okved_list(
    excel_text: str,
    llm_okved_list: list[str] | None = None,
    *,
    client: Any | None = None,
    model: str | None = None,
) -> list[str]:
    return resolve_okved_resolution(
        excel_text,
        llm_okved_list,
        client=client,
        model=model,
    ).codes


def resolve_okved_resolution(
    excel_text: str,
    llm_okved_list: list[str] | None = None,
    *,
    client: Any | None = None,
    model: str | None = None,
) -> OkvedResolution:
    """
    Resolve final OKVED codes using explicit codes, stop-rules, and local reference retrieval.
    Returns [] when the brief implies no industry restriction should be applied.
    """
    llm_okved_list = llm_okved_list or []
    try:
        reference = load_okved_reference(_load_reference_path())
    except FileNotFoundError:
        logger.warning("OKVED reference not found: %s", _load_reference_path())
        if _has_stop_signal(excel_text):
            return OkvedResolution(codes=[], explanations={}, decision_reason="no_filter_signal")
        codes = list(dict.fromkeys(_extract_explicit_codes(excel_text, llm_okved_list)))
        return OkvedResolution(
            codes=codes,
            explanations={code: ["Явно указан код ОКВЭД в брифе"] for code in codes},
            decision_reason="explicit_codes_without_reference",
        )

    explicit_codes = _filter_valid_codes(_extract_explicit_codes(excel_text, llm_okved_list), reference)
    if explicit_codes:
        return OkvedResolution(
            codes=explicit_codes,
            explanations={code: ["Явно указан код ОКВЭД в брифе"] for code in explicit_codes},
            decision_reason="explicit_codes",
        )

    if _has_stop_signal(excel_text):
        logger.info("[resolve_okved_list] Detected no-filter OKVED signal")
        return OkvedResolution(codes=[], explanations={}, decision_reason="no_filter_signal")

    context = _collect_target_context(excel_text)
    candidate_codes, candidate_explanations = _build_candidate_resolution(context, reference)
    if not candidate_codes:
        return OkvedResolution(codes=[], explanations={}, decision_reason="no_candidates")

    if client is not None and model and len(candidate_codes) > 1:
        reranked_codes = _llm_rerank_candidates(context, candidate_codes, reference, client, model)
        if reranked_codes is not None:
            explanations = {
                code: candidate_explanations.get(code, []) + ["Выбран LLM из кандидатов справочника"]
                for code in reranked_codes
            }
            return OkvedResolution(
                codes=reranked_codes,
                explanations=explanations,
                decision_reason="llm_rerank",
            )

    return OkvedResolution(
        codes=candidate_codes,
        explanations=candidate_explanations,
        decision_reason="local_candidates",
    )
