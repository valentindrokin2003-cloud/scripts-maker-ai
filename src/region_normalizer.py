import logging
import os
import re
from dataclasses import dataclass
from functools import lru_cache
from typing import List

import openpyxl

logger = logging.getLogger(__name__)

DEFAULT_OKATO_PATH = "data/okato.xlsx"

ALL_RUSSIA_INDICATORS = {
    "рф",
    "россия",
    "российская федерация",
    "all",
    "все",
    "russia",
    "russian federation",
}

ALL_RUSSIA_PATTERNS = (
    re.compile(r"\b(?:все|вся|весь)\s+(?:регионы|регионы рф|россия|территория россии)\b"),
    re.compile(r"\b(?:по|на)\s+всей\s+(?:рф|россии)\b"),
    re.compile(r"\bвсе\s+регионы\s+рф\b"),
)

REGION_MODIFIERS = {"область", "край", "округ", "республика", "регион"}

DISTRICT_FULL_NAMES = {
    "центральный": "ЦФО",
    "северо западный": "СЗФО",
    "северокавказский": "СКФО",
    "северо кавказский": "СКФО",
    "приволжский": "ПФО",
    "уральский": "УФО",
    "сибирский": "СФО",
    "дальневосточный": "ДФО",
    "южный": "ЮФО",
}

ADJECTIVE_ENDINGS = (
    "ская",
    "ский",
    "ской",
    "скои",
    "ское",
    "ская",
    "ая",
    "яя",
    "ий",
    "ый",
    "ой",
)

LEGACY_REGION_ALIASES = {
    "москва": "Москва",
    "санкт петербург": "Санкт-Петербург",
    "севастополь": "Севастополь",
    "амурская": "Амурская область",
    "архангельская": "Архангельская область",
    "астраханская": "Астраханская область",
    "белгородская": "Белгородская область",
    "брянская": "Брянская область",
    "владимирская": "Владимирская область",
    "волгоградская": "Волгоградская область",
    "вологодская": "Вологодская область",
    "воронежская": "Воронежская область",
    "еврейская": "Еврейская автономная область",
    "забайкальский": "Забайкальский край",
    "запорожская": "Запорожская область",
    "ивановская": "Ивановская область",
    "иркутская": "Иркутская область",
    "кабардино балкария": "Кабардино-Балкарская Республика",
    "калининград": "Калининградская область",
    "калининград область": "Калининградская область",
    "калмыкия": "Республика Калмыкия",
    "калужская": "Калужская область",
    "камчатка": "Камчатский край",
    "камчатский": "Камчатский край",
    "карачаево черкессия": "Карачаево-Черкесская Республика",
    "карелия": "Республика Карелия",
    "кемеровская": "Кемеровская область",
    "кировская": "Кировская область",
    "коми": "Республика Коми",
    "костромская": "Костромская область",
    "краснодар": "Краснодарский край",
    "краснодарский": "Краснодарский край",
    "красноярск": "Красноярский край",
    "красноярский": "Красноярский край",
    "крым": "Республика Крым",
    "курганская": "Курганская область",
    "курская": "Курская область",
    "ленинградская": "Ленинградская область",
    "липецкая": "Липецкая область",
    "магаданская": "Магаданская область",
    "марий эл": "Республика Марий Эл",
    "мордовия": "Республика Мордовия",
    "московская": "Московская область",
    "москва область": "Московская область",
    "мурманская": "Мурманская область",
    "нижегородская": "Нижегородская область",
    "новгородская": "Новгородская область",
    "новосибирск": "Новосибирская область",
    "новосибирская": "Новосибирская область",
    "оренбургская": "Оренбургская область",
    "орловская": "Орловская область",
    "осетия": "Республика Северная Осетия - Алания",
    "пензенская": "Пензенская область",
    "пермь": "Пермский край",
    "пермский": "Пермский край",
    "приморский": "Приморский край",
    "владивосток": "Приморский край",
    "псковская": "Псковская область",
    "республика алтай": "Республика Алтай",
    "алтай": "Республика Алтай",
    "республика башкортостан": "Республика Башкортостан",
    "башкортостан": "Республика Башкортостан",
    "башкирия": "Республика Башкортостан",
    "республика бурятия": "Республика Бурятия",
    "бурятия": "Республика Бурятия",
    "республика адыгея": "Республика Адыгея",
    "адыгея": "Республика Адыгея",
    "республика дагестан": "Республика Дагестан",
    "дагестан": "Республика Дагестан",
    "республика ингушетия": "Республика Ингушетия",
    "ингушетия": "Республика Ингушетия",
    "республика якутия": "Республика Саха (Якутия)",
    "якутия": "Республика Саха (Якутия)",
    "саха": "Республика Саха (Якутия)",
    "рязанская": "Рязанская область",
    "сахалинская": "Сахалинская область",
    "свердловская": "Свердловская область",
    "екатеринбург": "Свердловская область",
    "смоленская": "Смоленская область",
    "ставрополь": "Ставропольский край",
    "ставропольский": "Ставропольский край",
    "тамбовская": "Тамбовская область",
    "татарстан": "Республика Татарстан",
    "тверская": "Тверская область",
    "томск": "Томская область",
    "томская": "Томская область",
    "тульская": "Тульская область",
    "тюменская": "Тюменская область",
    "удмуртия": "Удмуртская Республика",
    "ульяновская": "Ульяновская область",
    "хабаровск": "Хабаровский край",
    "хабаровский": "Хабаровский край",
    "хакасия": "Республика Хакасия",
    "чеченская": "Чеченская Республика",
    "челябинск": "Челябинская область",
    "челябинская": "Челябинская область",
    "чувашия": "Чувашская Республика",
    "чукотский": "Чукотский автономный округ",
    "ямало ненецкий": "Ямало-Ненецкий автономный округ",
    "ярославская": "Ярославская область",
}


@dataclass(frozen=True)
class RegionFilters:
    regions: List[str]
    f_ocrygs: List[str]


def _normalize_key(value: str) -> str:
    normalized = value.casefold().replace("ё", "е")
    normalized = normalized.replace("«", " ").replace("»", " ")
    normalized = re.sub(r"[^a-zа-я0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def _clean_value(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).strip(" \"'«»")


def _is_all_russia_request(values: List[str]) -> bool:
    normalized_values = [_normalize_key(value) for value in values if _clean_value(value)]
    if any(value in ALL_RUSSIA_INDICATORS for value in normalized_values):
        return True

    joined = " ".join(normalized_values)
    return any(pattern.search(joined) for pattern in ALL_RUSSIA_PATTERNS)


def _split_region_parts(values: List[str]) -> list[str]:
    parts: list[str] = []
    for value in values:
        cleaned = _clean_value(value)
        if not cleaned:
            continue
        fragments = re.split(r"[,;\n]+|\s+\b(?:и|and)\b\s+", cleaned)
        parts.extend(fragment.strip() for fragment in fragments if fragment.strip())
    return parts


def _merge_modifier_fragments(values: list[str]) -> list[str]:
    merged: list[str] = []
    for value in values:
        normalized = _normalize_key(value)
        if normalized in REGION_MODIFIERS and merged:
            previous = merged.pop()
            merged.append(f"{previous} {value.strip()}")
        else:
            merged.append(value.strip())
    return merged


def _stem_location(value: str) -> str | None:
    token = _normalize_key(value).split(" ", 1)[0]
    replacements = {
        "ская": "ск",
        "ский": "ск",
        "ской": "ск",
        "скои": "ск",
        "ское": "ск",
        "ая": "",
        "яя": "",
        "ий": "",
        "ый": "",
        "ой": "",
    }
    for ending in ADJECTIVE_ENDINGS:
        if token.endswith(ending) and len(token) - len(ending) >= 4:
            return token[: -len(ending)] + replacements[ending]
    return None


def _district_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for base_name, short_name in DISTRICT_FULL_NAMES.items():
        for suffix in (
            "",
            " федеральный округ",
            " федральный округ",
            " федерельный округ",
            " федеральный",
            " федральный",
            " федерельный",
            " фо",
        ):
            aliases[f"{base_name}{suffix}".strip()] = short_name
    return aliases


@lru_cache(maxsize=4)
def _load_okato_reference(okato_path: str) -> tuple[dict[str, str], dict[str, str]]:
    workbook = openpyxl.load_workbook(okato_path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]

    try:
        region_index = header.index("region")
        district_index = header.index("f_ocryg")
    except ValueError as exc:
        raise ValueError(
            f"Missing required columns in okato reference: {okato_path}"
        ) from exc

    regions: dict[str, str] = {}
    districts: dict[str, str] = {}
    canonical_regions: set[str] = set()

    for row in rows:
        if region_index >= len(row) or district_index >= len(row):
            continue
        region = row[region_index]
        district = row[district_index]
        if not region or not district:
            continue

        region_name = str(region).strip()
        district_name = str(district).strip()
        canonical_regions.add(region_name)
        regions[_normalize_key(region_name)] = region_name
        districts[_normalize_key(district_name)] = district_name

    for alias, short_name in _district_aliases().items():
        districts[_normalize_key(alias)] = short_name

    for alias, canonical_name in LEGACY_REGION_ALIASES.items():
        canonical_key = _normalize_key(canonical_name)
        if canonical_key in regions:
            regions[_normalize_key(alias)] = regions[canonical_key]

    for canonical_name in canonical_regions:
        stem = _stem_location(canonical_name)
        if stem:
            regions.setdefault(stem, canonical_name)
            if stem.endswith("ск"):
                regions.setdefault(stem[:-2], canonical_name)

    return regions, districts


def _okato_path() -> str:
    return os.getenv("OKATO_PATH", DEFAULT_OKATO_PATH)


def normalize_region(region: str, followed_by_modifier: bool = False) -> str:
    del followed_by_modifier

    cleaned = _clean_value(region)
    if not cleaned:
        return cleaned

    try:
        region_aliases, _ = _load_okato_reference(_okato_path())
    except FileNotFoundError:
        logger.warning("OKATO reference not found: %s", _okato_path())
        return cleaned

    return region_aliases.get(_normalize_key(cleaned), cleaned)


def resolve_region_filters(values: List[str]) -> RegionFilters:
    """Split and normalize region filters into Spark `region` and `f_ocryg` lists."""
    if not values:
        return RegionFilters(regions=[], f_ocrygs=[])

    parts = _merge_modifier_fragments(_split_region_parts(values))
    if _is_all_russia_request(parts):
        logger.info("[resolve_region_filters] Detected all-regions indicator: %s", values)
        return RegionFilters(regions=[], f_ocrygs=[])

    try:
        region_aliases, district_aliases = _load_okato_reference(_okato_path())
    except FileNotFoundError:
        logger.warning("OKATO reference not found: %s", _okato_path())
        normalized_regions = [normalize_region(part) for part in parts if _clean_value(part)]
        deduplicated_regions = list(dict.fromkeys(normalized_regions))
        return RegionFilters(regions=deduplicated_regions, f_ocrygs=[])

    regions: list[str] = []
    f_ocrygs: list[str] = []
    seen_regions: set[str] = set()
    seen_f_ocrygs: set[str] = set()

    for part in parts:
        cleaned = _clean_value(part)
        if not cleaned:
            continue
        normalized = _normalize_key(cleaned)

        district = district_aliases.get(normalized)
        if district:
            if district not in seen_f_ocrygs:
                seen_f_ocrygs.add(district)
                f_ocrygs.append(district)
            continue

        region_name = region_aliases.get(normalized, cleaned)
        if region_name not in seen_regions:
            seen_regions.add(region_name)
            regions.append(region_name)

    return RegionFilters(regions=regions, f_ocrygs=f_ocrygs)


def normalize_regions(regions: List[str]) -> List[str]:
    """Normalize only subject-level regions for backward-compatible callers."""
    return resolve_region_filters(regions).regions
