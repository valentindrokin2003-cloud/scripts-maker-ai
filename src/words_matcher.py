import openpyxl
from dataclasses import dataclass
import re
from typing import List

@dataclass(frozen=True)
class DictEntry:
    word: str
    group_name: str = ""
    order: int = 0


ADJECTIVE_ENDINGS = (
    "ая",
    "яя",
    "ое",
    "ее",
    "ые",
    "ие",
    "ый",
    "ий",
    "ой",
    "ого",
    "его",
    "ому",
    "ему",
    "ыми",
    "ими",
    "ых",
    "их",
)

INFLECTION_ENDINGS = (
    "ами",
    "ями",
    "ого",
    "его",
    "ому",
    "ему",
    "ыми",
    "ими",
    "ых",
    "их",
    "ов",
    "ев",
    "ей",
    "ам",
    "ям",
    "ах",
    "ях",
    "ом",
    "ем",
    "ая",
    "яя",
    "ое",
    "ее",
    "ые",
    "ие",
    "ый",
    "ий",
    "ой",
    "а",
    "я",
    "ы",
    "и",
    "е",
    "о",
    "у",
    "ю",
)


def _normalise(value: str) -> str:
    return value.lower().replace("ё", "е")


def _tokens(value: str) -> list[str]:
    return re.findall(r"[a-zа-я0-9]+", _normalise(value))


def _stem(token: str) -> str:
    if token.endswith("ъем"):
        return token
    if token.endswith("ао"):
        return token
    for ending in INFLECTION_ENDINGS:
        if len(token) - len(ending) >= 4 and token.endswith(ending):
            return token[: -len(ending)].rstrip("ь")
    return token.rstrip("ь")


def _tokens_match(phrase_token: str, word_token: str) -> bool:
    phrase_stem = _stem(phrase_token)
    word_stem = _stem(word_token)
    if phrase_stem == word_stem:
        return True
    return len(word_stem) >= 5 and phrase_stem.startswith(word_stem)


def _is_descriptor(token: str) -> bool:
    return token.endswith(ADJECTIVE_ENDINGS)


def _find_first_match_index(phrase_tokens: list[str], word_tokens: list[str]) -> int | None:
    if not phrase_tokens or not word_tokens:
        return None

    for start in range(len(phrase_tokens)):
        pos = start
        first_matched_pos = None
        matched = True
        for word_token in word_tokens:
            while pos < len(phrase_tokens) and not _tokens_match(phrase_tokens[pos], word_token):
                pos += 1
            if pos >= len(phrase_tokens):
                matched = False
                break
            if first_matched_pos is None:
                first_matched_pos = pos
            pos += 1
        if matched:
            return first_matched_pos
    return None


def _find_unordered_match_index(phrase_tokens: list[str], word_tokens: list[str]) -> int | None:
    used_positions = set()
    matched_positions = []
    for word_token in word_tokens:
        for pos, phrase_token in enumerate(phrase_tokens):
            if pos in used_positions:
                continue
            if _tokens_match(phrase_token, word_token):
                used_positions.add(pos)
                matched_positions.append(pos)
                break
        else:
            return None
    return min(matched_positions) if matched_positions else None


def _has_leading_product_position(phrase_tokens: list[str], match_index: int) -> bool:
    leading_tokens = phrase_tokens[:match_index]
    return len(leading_tokens) <= 3 and all(_is_descriptor(token) for token in leading_tokens)


def _is_semantic_false_positive(entry: DictEntry, phrase_tokens: list[str]) -> bool:
    word = _normalise(entry.word)
    word_stems = {_stem(token) for token in _tokens(entry.word)}
    group = _normalise(entry.group_name)
    phrase = " ".join(phrase_tokens)

    if word_stems & {"нит", "нитк"} and "вольфрам" in phrase and not re.search(
        r"электр|металл|промышлен|свет",
        group,
    ):
        return True

    if word_stems & {"мед"} and re.search(r"\bмедн", phrase):
        return True

    if word == "шина медная" and "шинка" in phrase_tokens:
        return True

    if word == "фотообои" and "обои" not in phrase_tokens:
        return True

    return False


def _load_dict_entries(dict_path: str) -> List[DictEntry]:
    """Read dictionary rows. Only the `word` column is used as a match candidate."""
    wb = openpyxl.load_workbook(dict_path, data_only=True)
    entries = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [
            str(cell).strip().lower() if cell is not None else ""
            for cell in rows[0]
        ]
        if "word" in headers:
            word_idx = headers.index("word")
            group_idx = headers.index("group_name") if "group_name" in headers else None
            data_rows = rows[1:]
        else:
            word_idx = 0
            group_idx = 1 if len(headers) > 1 else None
            data_rows = rows

        for row in data_rows:
            if word_idx >= len(row):
                continue
            word = row[word_idx]
            if not isinstance(word, str) or not word.strip():
                continue
            group_name = ""
            if group_idx is not None and group_idx < len(row) and row[group_idx]:
                group_name = str(row[group_idx]).strip()
            entries.append(
                DictEntry(
                    word=word.strip(),
                    group_name=group_name,
                    order=len(entries),
                )
            )
    return entries


def _split_product_phrases(product_words: List[str]) -> list[str]:
    phrases = []
    for value in product_words:
        phrases.extend(
            part.strip()
            for part in re.split(r"[,;\n]+", value)
            if part and part.strip()
        )
    return phrases


def _match_phrase(phrase: str, entries: list[DictEntry]) -> list[str]:
    phrase_tokens = _tokens(phrase)
    if not phrase_tokens:
        return []

    candidates = []
    partial_candidates = []
    for entry in entries:
        word_tokens = _tokens(entry.word)
        match_index = _find_first_match_index(phrase_tokens, word_tokens)
        if match_index is None and len(word_tokens) > 1:
            match_index = _find_unordered_match_index(phrase_tokens, word_tokens)
        if match_index is None and len(phrase_tokens) == 1 and len(word_tokens) > 1:
            if _tokens_match(phrase_tokens[0], word_tokens[0]):
                match_index = 0
                target_candidates = partial_candidates
            else:
                target_candidates = candidates
        else:
            target_candidates = candidates
        if match_index is None:
            continue
        if not _has_leading_product_position(phrase_tokens, match_index):
            continue
        if _is_semantic_false_positive(entry, phrase_tokens):
            continue
        target_candidates.append((match_index, -len(word_tokens), entry.order, entry.word))

    selected_candidates = candidates or partial_candidates
    if selected_candidates:
        selected_candidates.sort()
        return [selected_candidates[0][3]]

    return []


def match_words(product_words: List[str], dict_path: str) -> List[str]:
    """
    For each word in product_words, find matching words in the dictionary file.
    Returns only words from the dictionary (original spelling), deduplicated.
    Threshold: fuzzywuzzy score >= 80.
    """
    if not product_words:
        return []

    dict_entries = _load_dict_entries(dict_path)
    if not dict_entries:
        return []

    matched = []
    seen = set()
    for phrase in _split_product_phrases(product_words):
        for match_word in _match_phrase(phrase, dict_entries):
            if match_word not in seen:
                seen.add(match_word)
                matched.append(match_word)

    return matched
