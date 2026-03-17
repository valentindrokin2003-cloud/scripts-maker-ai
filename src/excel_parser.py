import openpyxl


def parse_excel_to_text(path: str) -> str:
    """Read all sheets and cells from an Excel file into a single string."""
    wb = openpyxl.load_workbook(path, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Лист: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            row_str = " | ".join(v for v in row_values if v)
            if row_str:
                lines.append(row_str)
    return "\n".join(lines)
